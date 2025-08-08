import os
from openpyxl import load_workbook
from PIL import Image, ImageChops
from io import BytesIO

# === CONFIGURATION ===
excel_file = "happy_hour.xlsx"         # Path to Excel file
sheet_name = "REAL ONE"                     # Sheet to read
image_folder = "D:/bangkok_images"        # Folder with temp_image_*.png
image_columns = ['Image', 'Logo']           # Columns with embedded images
name_column = 'Name'                        # Column with name
id_column = 'happy_hours_id'                 # Column with happy_hour_id
output_csv = "matched_images_output.csv"    # Output CSV for matched results

# === Helper: Compare two images ===
def images_are_equal(img1, img2):
    if img1.size != img2.size:
        return False
    diff = ImageChops.difference(img1.convert("RGB"), img2.convert("RGB"))
    return not diff.getbbox()

# === Load workbook ===
wb = load_workbook(excel_file)
ws = wb[sheet_name]
headers = [cell.value for cell in ws[1]]
col_name_to_index = {col: idx + 1 for idx, col in enumerate(headers)}

# === Load temp images ===
temp_images = []
for fname in sorted(os.listdir(image_folder)):
    if fname.lower().endswith((".png", ".jpg", ".jpeg")):
        path = os.path.join(image_folder, fname)
        temp_images.append((fname, Image.open(path)))

# === Match and Output ===
matches = []

for img in ws._images:
    if hasattr(img.anchor, '_from'):
        anchor = img.anchor._from
        row_num = anchor.row + 1
        col_num = anchor.col + 1
        col_name = headers[col_num - 1]

        if col_name not in image_columns:
            continue

        # Get Name and happy_hour_id for that row
        name_value = ws.cell(row=row_num, column=col_name_to_index[name_column]).value
        id_value = ws.cell(row=row_num, column=col_name_to_index[id_column]).value

        # Get Excel embedded image
        excel_img_bytes = img._data()
        excel_pil_img = Image.open(BytesIO(excel_img_bytes)).convert("RGB")

        # Compare with all temp images
        for temp_fname, temp_pil_img in temp_images:
            if images_are_equal(excel_pil_img, temp_pil_img):
                matches.append({
                    "ExcelRow": row_num,
                    "happy_hours_id": id_value,
                    "Name": name_value,
                    "Column": col_name,
                    "MatchedFile": temp_fname
                })
                break

# === Print and optionally save ===
for match in matches:
    print(f"✅ Row {match['ExcelRow']} | ID: {match['happy_hours_id']} | Name: {match['Name']} | Column: {match['Column']} | File: {match['MatchedFile']}")

# Optional: Save to CSV
import csv
with open(output_csv, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=["ExcelRow", "happy_hours_id", "Name", "Column", "MatchedFile"])
    writer.writeheader()
    writer.writerows(matches)

print(f"\n🔍 Done. {len(matches)} images matched. Saved to {output_csv}")
